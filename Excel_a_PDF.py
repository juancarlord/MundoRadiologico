import datetime as dt
from sys import exit
from fpdf import FPDF
import openpyxl as xl
import xlrd
from tkinter import *
from tkinter import filedialog, ttk, messagebox
import os
import socket
import getpass
from uuid import getnode as getmac

MacAdd = getmac()
host = socket.gethostname()
ip = socket.gethostbyname(host)
userID = getpass.getuser()

filename = ''
month = dt.datetime.now()

mes = month.month
year = month.year
mesEs = {
    0: 'Diciembre',
    1: 'Enero',
    2: 'Febrero',
    3: 'Marzo',
    4: 'Abril',
    5: 'Mayo',
    6: 'Junio',
    7: 'Julio',
    8: 'Agosto',
    9: 'Septiembre',
    10: 'Octubre',
    11: 'Noviembre',
    12: 'Diciembre'
}
Tabs = {
    1: 'A',
    2: 'B',
    3: 'C',
    4: 'D',
    5: 'E',
    6: 'F',
    7: 'G',
    8: 'H',
    9: 'I',
    10: 'J',
    11: 'K',
    12: 'L',
    13: 'M',
    14: 'N',
    15: 'O',
    16: 'P',
    17: 'Q'
}
fecha = mesEs[mes-1].upper() + ' DE ' + str(year)

serial = {'Vj6k9B52ue',
          'kITXm80yWw',
          'lFKdHw60sA',
          'yq39lqC47e',
          'ppzhwv3xR8',
          'u0fAYnp4aL',
          'EBkz9p3QzC',
          '1KDlH7BOts',
          'oqPHIlSDVA',
          'tbEHmiZQLY'}


def free():
    if os.path.exists('redux.xlsx'):
        os.remove('redux.xlsx')
    else:
        messagebox.showinfo(
            'Atencion!', 'La memoria ya ha sido liberada', parent=window)
        print('El archivo no existe')


def letter_range(start, stop="{", step=1):
    """Yield a range of lowercase letters."""
    for ord_ in range(ord(start.upper()), ord(stop.upper()), step):
        yield chr(ord_)


def browseFiles():
    global filename
    filename = filedialog.askopenfilename(
        initialdir="/", title="Select a File", filetypes=(("Text files", "*.xlsx*"), ("all files", "*.*")))
    label_file_explorer.configure(text="Archivo Seleccionado: \n"+filename)
    book = xl.load_workbook(filename)
    label_sheets.configure(text="Escoja un libro \n de la lista de abajo")
    combo['values'] = (book.get_sheet_names())


def convert():
    hoja = combo.get()
    folder = fecha+'-'+hoja
    if os.path.exists(folder):
        print('La carpeta ya existe')
    else:
        os.mkdir(folder)
    # hoja = sheet_select.get().upper()

    wb = xl.load_workbook(filename)
    ws = wb.get_sheet_by_name(hoja)

    new = xl.Workbook()
    sheet = new.worksheets[0]

    mr = ws.max_row
    mc = ws.max_column

    print(mr)
    print(mc)
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            # reading cell value from source excel file
            c = ws.cell(row=i, column=j)

            # writing the read value to destination excel file
            sheet.cell(row=i, column=j).value = c.value

    new.save("redux.xlsx")

    filenameNew = "redux.xlsx"

    row = 2
    wb = xlrd.open_workbook(filenameNew, on_demand=True)
    sheet = wb.sheet_by_index(0)
    sheet.cell_value(0, 0)
    naRows = sheet.nrows
    naCol = sheet.ncols+1
    print(sheet.ncols)
    print(sheet.nrows)
    finLetra = Tabs[naCol]
    print(naRows)
    print(finLetra)
    nTitulo = 2
    loops = 100/naRows
    progress['value'] = loops
    wb.release_resources()
    while row <= naRows:
        row += 1

        class PDF(FPDF):
            def header(self):
                self.image('data/logo.png', 50, 5, 40, 20)
                self.set_font('Arial', '', 12)
                self.cell(110)
                self.cell(30, 10, 'MUNDO RADIOLOGICO S.A.S.')
                self.ln()
                self.cell(120)
                self.cell(30, 10, str(fecha))
                self.ln(20)

        pdf = PDF('L', 'mm', 'Legal')
        pdf.add_page()
        pdf.set_font('Arial', 'b', 6)
        pdf.set_title('title')

        book = xl.load_workbook(filenameNew)
        sheet = book.worksheets[0]
        conTitulo = 'A' + str(nTitulo)

        title = sheet[str(conTitulo)]

        for i in list(letter_range('a', finLetra)):
            num = 1
            valor = i + str(num)
            nextTitle = sheet[valor]
            name = nextTitle.value
            pdf.cell(24, 5, str(name), 1, 0, 'C')
        pdf.ln()
        sTitulo = nTitulo

        nextCell = 'A' + str(sTitulo)
        nextData = sheet[str(nextCell)]
        nciclos = nTitulo
        if title.value != None:
            if title.value == nextData.value:
                num1 = title.value
                num2 = nextData.value
                reps = 1
                while num1 == num2:
                    if reps > 8:
                        pdf.add_page()
                        pdf.set_font('Arial', 'b', 6)
                        for i in list(letter_range('a', finLetra)):
                            num = 1
                            valor = i + str(num)
                            nextTitle = sheet[valor]
                            name = nextTitle.value
                            pdf.cell(24, 5, str(name), 1, 0, 'C')
                        pdf.ln()
                        reps = 1
                    pdf.set_font('Arial', '', 5)
                    for i in list(letter_range('a', finLetra)):
                        valor2 = i + str(nciclos)
                        nextData = sheet[valor2]
                        name2 = nextData.value
                        x = pdf.get_x()
                        y = pdf.get_y()
                        pdf.multi_cell(24, 3, str(name2), 0, 'C')
                        pdf.set_xy(x+24, y)
                    pdf.ln(20)
                    sTitulo += 1
                    nciclos += 1
                    nextCell = 'A' + str(sTitulo)
                    nextData = sheet[str(nextCell)]
                    num2 = nextData.value
                    reps += 1
                try:
                    nombre = title.value+'.pdf'
                except TypeError:
                    print('Se acabo el index')
                    break
                else:
                    print('Se creo el documento')
                nTitulo += 1
                pdf.output(folder+'/'+nombre, 'F')
                print(title.value)
                progress['value'] += loops
                if num1 != num2:
                    conTitulo = nextCell
                    nTitulo = sTitulo
                    title = sheet[str(conTitulo)]
                    row = sTitulo
    messagebox.showinfo(
        'Importante', 'Conversion completa! \nPresione liberar', parent=window)


def validate():
    if contPass.get() == 'Vj6k9B52ue':
        if str(MacAdd) != '49410867780297':
            messagebox.showerror(
                'Equipo no valido', 'Este equipo no esta autorizado para el uso de esta aplicacion')
            exit()
        cont.destroy()
        window.deiconify()
    else:
        messagebox.showerror(
            'Sin licencia', 'Ingrese el codigo de serie correcto de lo contrario\npongase en contacto con soporte')
        exit()


window = Tk()
window.title('Generador de Excel a PDF')

window.config()
window.iconbitmap(r'data/favicon.ico')

img = PhotoImage(file='data/logo1.png')

photolabel = Label(window, image=img)

label_file_explorer = Label(
    window, text="Buscar archivo en Excel para la conversion", width=50, height=4, fg="blue")
label_sheets = Label(
    window, text="Cuando cargue el archivo use la lista plegable \n para seleccionar el libro", width=50, height=4, fg='blue')
sheet_select = Entry(window, font=('Arial', 12, 'normal'))
button_explore = Button(window, text="Buscar", command=browseFiles)
button_exit = Button(window,  text="Salir", command=exit)
button_start = Button(window, text="Ejecutar", command=convert)
progress = ttk.Progressbar(window, orient=HORIZONTAL,
                           length=200, mode='determinate')
n = StringVar()
combo = ttk.Combobox(window, width=27, textvariable=n)
liberar = Button(window, text='Liberar', command=free)
label_file_explorer.grid(column=1, row=1)
button_explore.grid(column=1, row=2)
label_sheets.grid(column=1, row=3)
# sheet_select.grid(column=1, row=4)
combo.grid(column=1, row=5)
button_start.grid(column=1, row=6, padx=10, pady=10)
# progress.grid(column=1, row=7)
liberar.grid(column=1, row=8)
button_exit.grid(column=1, row=9, padx=10, pady=10)
photolabel.grid(column=1, row=10)

windowWidth = window.winfo_reqwidth()
windowHeight = window.winfo_reqheight()
print(windowHeight, ' ', windowWidth)
positionRight = int(window.winfo_screenwidth()/2 - windowWidth)
positionDown = int(window.winfo_screenheight()/2 - windowHeight)

window.geometry("+{}+{}".format(positionRight, positionDown))

window.withdraw()


cont = Tk()
cont.title('Verificador')
cont.config()
contLabel = Label(cont, text='Ingrese el serial del programa',
                  width=30, height=4)
contPass = Entry(cont, show='*')
contBtn = Button(cont, text='Validar', command=validate)
contLabel.grid(column=1, row=1)
contPass.grid(column=1, row=2)
contBtn.grid(column=1, row=3, padx=10, pady=10)
cont.mainloop()
